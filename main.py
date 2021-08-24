# main libs
from dotenv import load_dotenv
from os import getenv
from typing import List
import tkinter as tk
from tkinter import filedialog
from openpyxl import load_workbook
from openpyxl.cell import Cell

# local
from project_types import Student
from sheets_service import Worksheet
from zipfile import ZipFile

load_dotenv()


def as_student(section: int, row: List[Cell]) -> Student:
    (last_name, second_last_name, first_name, rut, student_number, curriculum, email) = row
    return (
        str(section),
        student_number.value or "",
        last_name.value or "",
        second_last_name.value or "",
        first_name.value or "",
        email.value or "",
        rut.value or "",
        curriculum.value or "",
    )


def main():
    root = tk.Tk()
    root.withdraw()

    file_path = filedialog.askopenfilename(filetypes=[["Zip file", "zip"]])
    course_students = set()
    with ZipFile(file_path) as zipfile:
        files = filter(lambda f: "__MACOSX" not in f, zipfile.namelist())
        for section_path in files:
            with zipfile.open(section_path, mode="r") as section_xslx:
                section_number = int(section_path.split(" - ")[0].split("_")[-1])
                wb = load_workbook(section_xslx)
                ws = wb.active
                _, end = ws.calculate_dimension().split(":")
                student_list = set(
                    map(
                        lambda row, number=section_number: as_student(number, row), ws[f"B8:{end}"]
                    )
                )
                course_students |= student_list
    gs = Worksheet(sheet_id=getenv("SHEET_ID"))
    gs.add_students(course_students - gs.current_students)


if __name__ == "__main__":
    main()
